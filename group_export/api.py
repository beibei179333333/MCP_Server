"""HTTP client for the fun-stat-bot API.

Because the exact schema can change, this client *auto-discovers* the export
endpoint from the Swagger/OpenAPI spec and tolerates several pagination and
response-envelope conventions. Everything it auto-detects can be overridden
explicitly (see CLI flags / ApiConfig).
"""
from __future__ import annotations

import csv
import io
import time
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import requests


SWAGGER_CANDIDATES = [
    "/swagger/v1/swagger.json",
    "/swagger/v1/swagger.yaml",
    "/swagger.json",
    "/openapi.json",
    "/v1/swagger.json",
    "/api/swagger.json",
    "/api-docs",
    "/swagger/docs/v1",
]

# Keywords used to score how likely a path is the "export group members" op.
PATH_KEYWORDS = ["member", "participant", "成员", "群成员", "groupuser", "group_user"]
EXPORT_KEYWORDS = ["export", "导出", "download", "list", "all"]

# Response envelope keys that may wrap the member array.
ENVELOPE_KEYS = ["data", "items", "result", "results", "list", "members",
                 "rows", "records", "users", "participants", "content"]
# Pagination "total count" keys.
TOTAL_KEYS = ["total", "totalCount", "total_count", "count", "totalRecords"]


@dataclass
class ApiConfig:
    base_url: str
    token: str
    # Optional explicit overrides (skip auto-discovery when set).
    endpoint: Optional[str] = None          # e.g. "/api/group/members/export"
    method: str = "GET"
    group_param: Optional[str] = None        # query/body param carrying the group id
    page_param: Optional[str] = None         # e.g. "page" / "offset"
    size_param: Optional[str] = None         # e.g. "page_size" / "limit"
    page_size: int = 200
    page_starts_at: int = 1                  # 1 for page-based, 0 for offset-based
    paginate_by_offset: bool = False         # True => page_param is an item offset
    extra_params: Dict[str, Any] = field(default_factory=dict)
    timeout: int = 60
    max_retries: int = 4
    retry_backoff: float = 2.0
    sleep_between_pages: float = 0.3
    verbose: bool = True


class ApiClient:
    def __init__(self, cfg: ApiConfig):
        self.cfg = cfg
        self.base_url = cfg.base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {cfg.token}",
            "Accept": "application/json",
            "User-Agent": "group-export/1.0",
        })
        self.spec: Optional[dict] = None

    # ---- low-level request with retry/backoff -------------------------------
    def _request(self, method: str, url: str, **kw) -> requests.Response:
        delay = self.cfg.retry_backoff
        last_exc: Optional[Exception] = None
        for attempt in range(1, self.cfg.max_retries + 1):
            try:
                resp = self.session.request(
                    method, url, timeout=self.cfg.timeout, **kw
                )
                if resp.status_code in (429, 500, 502, 503, 504):
                    raise requests.HTTPError(f"retryable status {resp.status_code}")
                return resp
            except (requests.RequestException, ) as exc:
                last_exc = exc
                if attempt == self.cfg.max_retries:
                    break
                self._log(f"  request failed ({exc}); retry {attempt} in {delay:.0f}s")
                time.sleep(delay)
                delay *= 2
        raise RuntimeError(f"request to {url} failed after retries: {last_exc}")

    def _log(self, msg: str) -> None:
        if self.cfg.verbose:
            print(msg, flush=True)

    # ---- swagger discovery --------------------------------------------------
    def fetch_spec(self) -> Optional[dict]:
        if self.spec is not None:
            return self.spec
        for path in SWAGGER_CANDIDATES:
            url = self.base_url + path
            try:
                resp = self._request("GET", url)
            except RuntimeError:
                continue
            if resp.status_code != 200:
                continue
            try:
                data = resp.json()
            except ValueError:
                continue
            if isinstance(data, dict) and ("paths" in data or "swagger" in data
                                           or "openapi" in data):
                self._log(f"discovered swagger spec at {path}")
                self.spec = data
                return data
        self._log("could not auto-fetch swagger spec (continuing with overrides)")
        return None

    def discover_endpoint(self) -> Optional[Tuple[str, str, dict]]:
        """Return (path, method, operation) best matching 'export members'."""
        spec = self.fetch_spec()
        if not spec or "paths" not in spec:
            return None
        best = None
        best_score = -1
        for path, ops in spec["paths"].items():
            if not isinstance(ops, dict):
                continue
            for method, op in ops.items():
                if method.lower() not in ("get", "post"):
                    continue
                hay = " ".join(str(x) for x in (
                    path, op.get("operationId", ""), op.get("summary", ""),
                    op.get("description", ""), " ".join(op.get("tags", []) or []),
                )).lower()
                score = 0
                if any(k in hay for k in PATH_KEYWORDS):
                    score += 3
                if any(k in hay for k in EXPORT_KEYWORDS):
                    score += 2
                if "group" in hay or "群" in hay or "chat" in hay:
                    score += 1
                if score > best_score:
                    best_score = score
                    best = (path, method.upper(), op or {})
        if best and best_score > 0:
            self._log(f"selected endpoint: {best[1]} {best[0]} (score {best_score})")
            return best
        return None

    @staticmethod
    def _detect_pagination(op: dict) -> Dict[str, Any]:
        """Infer page/size/offset param names from an operation's parameters."""
        params = op.get("parameters", []) or []
        names = {p.get("name", "").lower(): p.get("name") for p in params
                 if isinstance(p, dict)}
        out: Dict[str, Any] = {}
        for cand in ("page", "pageindex", "page_index", "pagenumber", "page_number"):
            if cand in names:
                out["page_param"] = names[cand]
                out["paginate_by_offset"] = False
                out["page_starts_at"] = 1
                break
        if "page_param" not in out:
            for cand in ("offset", "skip", "start"):
                if cand in names:
                    out["page_param"] = names[cand]
                    out["paginate_by_offset"] = True
                    out["page_starts_at"] = 0
                    break
        for cand in ("page_size", "pagesize", "size", "limit", "count", "per_page",
                     "perpage", "take"):
            if cand in names:
                out["size_param"] = names[cand]
                break
        for cand in ("group_id", "groupid", "chat_id", "chatid", "group", "chat",
                     "id", "gid"):
            if cand in names:
                out["group_param"] = names[cand]
                break
        return out

    def resolve(self) -> ApiConfig:
        """Fill any unset endpoint/pagination fields via discovery."""
        cfg = self.cfg
        if cfg.endpoint:
            return cfg
        found = self.discover_endpoint()
        if found:
            path, method, op = found
            cfg.endpoint = path
            cfg.method = method
            detected = self._detect_pagination(op)
            for k, v in detected.items():
                if getattr(cfg, k, None) in (None, ) or (
                    k == "page_param" and not cfg.page_param
                ) or (k == "size_param" and not cfg.size_param) or (
                    k == "group_param" and not cfg.group_param
                ):
                    setattr(cfg, k, v)
        return cfg

    # ---- response parsing ---------------------------------------------------
    @staticmethod
    def _extract_list(payload: Any) -> Tuple[List[dict], Optional[int]]:
        """Find the member array and optional total count in a JSON payload."""
        total = None
        if isinstance(payload, list):
            return [x for x in payload if isinstance(x, dict)], None
        if isinstance(payload, dict):
            for tk in TOTAL_KEYS:
                if isinstance(payload.get(tk), int):
                    total = payload[tk]
                    break
            for key in ENVELOPE_KEYS:
                val = payload.get(key)
                if isinstance(val, list):
                    return [x for x in val if isinstance(x, dict)], total
                if isinstance(val, dict):
                    inner, t2 = ApiClient._extract_list(val)
                    if inner:
                        return inner, total if total is not None else t2
            # Fallback: first list-of-dicts anywhere in the dict.
            for val in payload.values():
                if isinstance(val, list) and val and isinstance(val[0], dict):
                    return val, total
        return [], total

    @staticmethod
    def _parse_csv_bytes(content: bytes) -> List[dict]:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(r) for r in reader]

    def _parse_response(self, resp: requests.Response) -> Tuple[List[dict], Optional[int]]:
        ctype = resp.headers.get("Content-Type", "").lower()
        if "json" in ctype:
            return self._extract_list(resp.json())
        if "csv" in ctype or "text/plain" in ctype:
            return self._parse_csv_bytes(resp.content), None
        if "spreadsheet" in ctype or "excel" in ctype or ctype.endswith("xlsx"):
            return self._parse_xlsx_bytes(resp.content), None
        # Try JSON regardless; many servers mislabel content-type.
        try:
            return self._extract_list(resp.json())
        except ValueError:
            return self._parse_csv_bytes(resp.content), None

    @staticmethod
    def _parse_xlsx_bytes(content: bytes) -> List[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("xlsx response needs openpyxl installed") from exc
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h) for h in next(rows)]
        out = []
        for r in rows:
            out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
        return out

    # ---- the main fetch loop ------------------------------------------------
    def fetch_members(self, group: str) -> List[dict]:
        cfg = self.resolve()
        if not cfg.endpoint:
            raise RuntimeError(
                "no export endpoint resolved. Pass --endpoint explicitly "
                "(see what your swagger UI shows for 'export group members')."
            )
        url = self.base_url + cfg.endpoint
        results: List[dict] = []
        seen_pages_empty = 0
        page = cfg.page_starts_at
        total: Optional[int] = None
        self._log(f"fetching members for group={group} via {cfg.method} {cfg.endpoint}")

        while True:
            params = dict(cfg.extra_params)
            if cfg.group_param:
                params[cfg.group_param] = group
            else:
                # Best effort if discovery missed the name.
                params.setdefault("group_id", group)
            if cfg.size_param:
                params[cfg.size_param] = cfg.page_size
            if cfg.page_param:
                params[cfg.page_param] = page

            if cfg.method.upper() == "GET":
                resp = self._request("GET", url, params=params)
            else:
                resp = self._request("POST", url, json=params)

            if resp.status_code == 401:
                raise RuntimeError("401 Unauthorized — token invalid or expired.")
            if resp.status_code == 404:
                raise RuntimeError(f"404 for {url} — wrong endpoint? {resp.text[:200]}")
            if resp.status_code >= 400:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:300]}")

            batch, t = self._parse_response(resp)
            if t is not None:
                total = t
            results.extend(batch)
            self._log(f"  page {page}: +{len(batch)} (running total {len(results)}"
                      + (f" / {total}" if total else "") + ")")

            # Stop conditions.
            if not cfg.page_param:
                break  # endpoint isn't paginated; one shot.
            if not batch:
                seen_pages_empty += 1
                if seen_pages_empty >= 1:
                    break
            if total is not None and len(results) >= total:
                break
            if len(batch) < cfg.page_size and cfg.size_param:
                break  # short page => last page

            page += cfg.page_size if cfg.paginate_by_offset else 1
            time.sleep(cfg.sleep_between_pages)

        self._log(f"fetch complete: {len(results)} raw records")
        return results
